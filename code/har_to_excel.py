import json, base64, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

har_file = sys.argv[1] if len(sys.argv) > 1 else "www_realestate_com_au.har"
out_file = sys.argv[2] if len(sys.argv) > 2 else "rea_sold_listings.xlsx"

with open(har_file, "r") as f:
    har = json.load(f)

entries = har["log"]["entries"]
lexa_entries = [
    e for e in entries
    if "lexa.realestate.com.au/graphql" in e["request"]["url"]
    and e["request"]["method"] == "POST"
    and e["response"]["status"] == 200
]

print(f"Found {len(lexa_entries)} lexa GraphQL responses")

all_listings = []
for entry in lexa_entries:
    content = entry["response"]["content"]
    text = content.get("text", "")
    if content.get("encoding") == "base64":
        text = base64.b64decode(text).decode("utf-8")
    try:
        data = json.loads(text)
        results = data["data"]["soldSearch"]["results"]
        page = results["pagination"]["page"]
        exact = results["exact"].get("items", [])
        surrounding = (results.get("surrounding") or {}).get("items", [])
        items = exact + surrounding
        print(f"  Page {page}: {len(exact)} exact + {len(surrounding)} surrounding = {len(items)} listings")
        for item in items:
            l = item.get("listing", {})
            address = l.get("address", {})
            price = l.get("price", {}) or {}
            features = l.get("generalFeatures", {}) or {}
            sizes = l.get("propertySizes", {}) or {}
            prop_type = l.get("propertyType", {}) or {}
            date_sold = l.get("dateSold", {}) or {}
            land = sizes.get("land", {}) or {}
            building = sizes.get("building", {}) or {}
            all_listings.append({
                "id": l.get("id"),
                "full_address": (address.get("display") or {}).get("fullAddress"),
                "street": (address.get("display") or {}).get("shortAddress"),
                "suburb": address.get("suburb"),
                "state": address.get("state"),
                "postcode": address.get("postcode"),
                "sold_price": price.get("display"),
                "date_sold": date_sold.get("display"),
                "property_type": prop_type.get("display"),
                "bedrooms": (features.get("bedrooms") or {}).get("value"),
                "bathrooms": (features.get("bathrooms") or {}).get("value"),
                "car_spaces": (features.get("parkingSpaces") or {}).get("value"),
                "land_size_m2": land.get("displayValue"),
                "building_size_m2": building.get("displayValue") if building else None,
            })
    except Exception as ex:
        print(f"  Skipped entry: {ex}")

print(f"Total listings: {len(all_listings)}")

wb = Workbook()
ws = wb.active
ws.title = "Sold Listings"

headers = ["ID", "Full Address", "Street", "Suburb", "State", "Postcode",
           "Sold Price", "Date Sold", "Property Type", "Bedrooms", "Bathrooms",
           "Car Spaces", "Land Size (m²)", "Building Size (m²)"]

header_fill = PatternFill("solid", start_color="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Arial")
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

keys = ["id", "full_address", "street", "suburb", "state", "postcode",
        "sold_price", "date_sold", "property_type", "bedrooms", "bathrooms",
        "car_spaces", "land_size_m2", "building_size_m2"]

for row, listing in enumerate(all_listings, 2):
    for col, key in enumerate(keys, 1):
        ws.cell(row=row, column=col, value=listing.get(key))

col_widths = [12, 45, 30, 20, 8, 10, 14, 14, 16, 10, 10, 12, 15, 18]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(out_file)
print(f"Saved to {out_file}")
